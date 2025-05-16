from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
import inspect

from config import UPDATED_FILE, FILTERED_FILE
from logger_config import setup_logger

logger = setup_logger(__name__)

def log_variables(local_vars: Dict[str, Any], exclude: List[str] = None) -> None:
    """Log variable names and their values"""
    if exclude is None:
        exclude = []
    exclude.extend(['self', 'args', 'kwargs', 'exclude'])
    
    frame = inspect.currentframe().f_back
    try:
        for var_name, var_value in frame.f_locals.items():
            if var_name not in exclude:
                logger.debug(f"Variable: {var_name} = {var_value!r}")
    finally:
        del frame

def filter_columns() -> None:
    logger.info("Filtering columns to create filtered sheet")
    log_variables(locals())
    
    try:
        logger.info(f"Loading workbook from {UPDATED_FILE}")
        wb = load_workbook(filename=UPDATED_FILE)
        ws = wb.active
        logger.info(f"Active worksheet: {ws.title}, Rows: {ws.max_row}, Columns: {ws.max_column}")
        
        logger.info("Creating 'Filtered' worksheet")
        filtered_ws = wb.create_sheet(title="Filtered")
        
        columns_to_keep = [
            "columna en informe diario", "Hora de Análisis", "Saturación (%) (Pureza)",
            "Longitud de onda (nm)", "L*", "a*", "b*", "Densidad", "% T 550 (2mm)",
            "Semillas L593", "Semillas L594", "Semillas (0 - 0,5) mm L 593",
            "Semillas (0 - 0,5) mm L 594", "Burbujas (0,5-1) mm L 593",
            "Burbujas (0,5-1) mm L 594", "Burbujas (>1)mm L 593", "Burbujas (>1)mm L 594",
            "Burbujas por Kg - 593", "Burbujas por Kg - 594", "SiO2", "Na2O", "CaO",
            "MgO", "Al2O3", "K2O", "SO3", "Fe2O3", "TiO2", "SiO2D (100-S(ox))",
            "Cr2O3", "%FeO as Fe2O3", "Redox", "Viscosidad (°C)", "Cooling Time (s)"
        ]
        logger.debug(f"Columns to keep: {columns_to_keep}")
        
        columns_found = 0
        columns_not_found = []
        
        for col_name in columns_to_keep:
            col_idx = None
            logger.debug(f"Looking for column: {col_name}")
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value and str(cell_value).strip() == col_name:
                    col_idx = col
                    logger.debug(f"Found '{col_name}' at column {get_column_letter(col_idx)}")
                    break
                    
            if col_idx is not None:
                dest_col = columns_to_keep.index(col_name) + 1
                logger.debug(f"Copying column {get_column_letter(col_idx)} to column {get_column_letter(dest_col)} in filtered sheet")
                
                rows_copied = 0
                for row in range(1, ws.max_row + 1):
                    source_cell = ws.cell(row=row, column=col_idx)
                    filtered_ws.cell(row=row, column=dest_col, value=source_cell.value)
                    rows_copied += 1
                
                logger.debug(f"Copied {rows_copied} rows from column {get_column_letter(col_idx)} to filtered sheet")
                columns_found += 1
            else:
                logger.warning(f"Column not found: '{col_name}'")
                columns_not_found.append(col_name)
        
        logger.info(f"Filtering complete. Found {columns_found} of {len(columns_to_keep)} columns.")
        if columns_not_found:
            logger.warning(f"Columns not found: {', '.join(columns_not_found)}")
        
        logger.info(f"Saving filtered workbook to {FILTERED_FILE}")
        wb.save(FILTERED_FILE)
        logger.info(f"Successfully saved filtered sheet to: {FILTERED_FILE}")
        
    except Exception as e:
        logger.error(f"Error in filter_columns: {str(e)}", exc_info=True)
        raise
