from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from typing import List, Any, Dict
from datetime import datetime, time
import inspect

from config import ORIGINAL_FILE, UPDATED_FILE
from logger_config import setup_logger
from processor.utils import parse_time_string, process_time_cells

logger = setup_logger(__name__)

def log_variables(local_vars: Dict[str, Any], exclude: List[str] = None) -> None:
    """Log variable names and their values"""
    if exclude is None:
        exclude = []
    exclude.extend(['self', 'args', 'kwargs', 'exclude'])
    
    frame = inspect.currentframe().f_back
    try:
        for var_name, var_value in frame.f_locals.items():
            if var_name not in exclude:
                logger.debug(f"Variable: {var_name} = {var_value!r}")
    finally:
        del frame

def unmerge_columns(ws) -> None:
    logger.info("Unmerging cells in range B26 to E31...")
    start_row, end_row = 26, 31
    start_col, end_col = 2, 5
    log_variables(locals(), ['ws'])

    merged_ranges = list(ws.merged_cells.ranges)
    logger.debug(f"Found {len(merged_ranges)} merged ranges in the worksheet")
    
    ranges_unmerged = 0
    for merged_range in merged_ranges:
        if (start_row <= merged_range.min_row <= end_row and
            start_row <= merged_range.max_row <= end_row and
            start_col <= merged_range.min_col <= end_col and
            start_col <= merged_range.max_col <= end_col):
            logger.debug(f"Unmerging range: {merged_range}")
            ws.unmerge_cells(str(merged_range))
            ranges_unmerged += 1
    logger.info(f"Unmerged {ranges_unmerged} ranges in B26:E31")

    logger.info("Unmerging any remaining merged cells in columns A-G...")
    ranges_unmerged = 0
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_col <= 7:
            logger.debug(f"Unmerging range in columns A-G: {merged_range}")
            ws.unmerge_cells(str(merged_range))
            ranges_unmerged += 1
    logger.info(f"Unmerged {ranges_unmerged} additional ranges in columns A-G")

def set_row_values(ws, values: List[str], start_row: int = 26) -> None:
    logger.info(f"Setting values from row {start_row}")
    log_variables(locals(), ['ws'])
    
    for i, value in enumerate(values, start=start_row):
        cell = ws.cell(row=i, column=2)
        logger.debug(f"Processing row {i}, value: {value}")
        
        if isinstance(cell, MergedCell):
            logger.debug(f"Cell {cell.coordinate} is part of a merged range")
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    logger.debug(f"Found merged range: {merged_range}")
                    target_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                    logger.debug(f"Setting value in merged cell at {target_cell.coordinate}")
                    target_cell.value = value
                    logger.info(f"Set merged cell {target_cell.coordinate} to: {value}")
                    break
        else:
            logger.debug(f"Setting value in regular cell {cell.coordinate}")
            cell.value = value
            logger.info(f"Set cell {cell.coordinate} to: {value}")

def move_hora_values(ws) -> None:
    logger.info("Moving 'hora' values from column F to column B...")
    log_variables(locals(), ['ws'])
    
    rows_processed = 0
    for row in range(1, ws.max_row + 1):
        cell_f = ws.cell(row=row, column=6)
        cell_b = ws.cell(row=row, column=2)
        
        if cell_f.value and "hora" in str(cell_f.value).lower():
            logger.debug(f"Found 'hora' value at F{row}: {cell_f.value}")
            
            if isinstance(cell_b, MergedCell):
                logger.debug(f"Target cell B{row} is part of a merged range")
                for merged_range in ws.merged_cells.ranges:
                    if cell_b.coordinate in merged_range:
                        target_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                        logger.debug(f"Setting value in merged cell {target_cell.coordinate}")
                        target_cell.value = cell_f.value
                        logger.info(f"Moved 'hora' value from F{row} to merged cell {target_cell.coordinate}")
                        break
            else:
                logger.debug(f"Setting value in regular cell B{row}")
                cell_b.value = cell_f.value
                logger.info(f"Moved 'hora' value from F{row} to B{row}")
                
            cell_f.value = None
            rows_processed += 1
    
    logger.info(f"Moved 'hora' values in {rows_processed} rows")

def process_dates(ws, rows_to_update: List[int]) -> None:
    """
    Process date and time values in the worksheet.
    
    Args:
        ws: Worksheet to process
        rows_to_update: List of row indices to process (1-based)
    """
    logger.info("Processing date/time values from columns G to HH...")
    log_variables(locals())
    
    date_columns = []
    date_objs = {}
    
    # First pass: identify date columns and parse dates
    for col in range(7, 209):  # Columns G to HH
        date_cell = ws.cell(row=1, column=col)
        if date_cell.value is None:
            continue
            
        try:
            if isinstance(date_cell.value, str):
                try:
                    # Parse date string (e.g., "25-Apr-25")
                    day, month, year = date_cell.value.split('-')
                    year = f"20{year}"  # Assuming 2-digit year
                    date_str = f"{day}-{month}-{year}"
                    date_obj = datetime.strptime(date_str, '%d-%b-%Y')
                except ValueError:
                    logger.debug(f"Skipping non-date column {get_column_letter(col)}: {date_cell.value}")
                    continue
            else:
                # Already a date object
                date_obj = date_cell.value
                
            date_columns.append(col)
            date_objs[col] = date_obj.date()  # Store date part only
            
        except Exception as e:
            logger.error(f"Error processing date in column {get_column_letter(col)}: {e}")
            continue
    
    logger.info(f"Found {len(date_columns)} date columns to process")
    
    # Process time values in each date column
    for col in date_columns:
        col_letter = get_column_letter(col)
        date_str = date_objs[col].strftime('%Y-%m-%d')
        
        logger.info(f"Processing time values in column {col_letter} for date {date_str}")
        
        # Process time values in this column
        time_results = process_time_cells(ws, [col], rows_to_update)
        
        # Combine dates with times
        updates = 0
        for row in rows_to_update:
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, time):
                combined_dt = datetime.combine(date_objs[col], cell.value)
                cell.value = combined_dt
                updates += 1
        
        logger.info(f"Updated {updates} datetime values in column {col_letter}")
    
    logger.info(f"Completed processing of {len(date_columns)} date columns")

def set_column_widths(ws) -> None:
    logger.info("Setting column widths from C to HH...")
    log_variables(locals(), ['ws'])
    
    columns_updated = 0
    for col in range(3, 209):
        col_letter = get_column_letter(col)
        try:
            ws.column_dimensions[col_letter].width = 15
            columns_updated += 1
        except Exception as e:
            logger.error(f"Error setting width for column {col_letter}: {e}")
    
    logger.info(f"Updated width for {columns_updated} columns")

def clean_daily_excel() -> None:
    logger.info(f"Cleaning Excel file: {ORIGINAL_FILE}")
    log_variables(locals())
    
    try:
        logger.debug(f"Loading workbook from {ORIGINAL_FILE}")
        wb = load_workbook(filename=ORIGINAL_FILE)
        ws = wb.active
        logger.info(f"Active worksheet: {ws.title}")

        values_to_set = [
            "Semillas L593", "Semillas L594", "Semillas (0 - 0,5) mm L 593",
            "Semillas (0 - 0,5) mm L 594", "Burbujas (0,5-1) mm L 593",
            "Burbujas (0,5-1) mm L 594", "Burbujas (>1)mm L 593", "Burbujas (>1)mm L 594"
        ]
        logger.debug(f"Values to set: {values_to_set}")

        logger.info("Starting worksheet processing...")
        unmerge_columns(ws)
        set_row_values(ws, values_to_set)
        move_hora_values(ws)
        
        rows_to_update = list(range(26, 34))
        logger.debug(f"Will update date/time in rows: {rows_to_update}")
        process_dates(ws, rows_to_update)
        
        set_column_widths(ws)

        logger.info(f"Saving cleaned workbook to {UPDATED_FILE}")
        wb.save(UPDATED_FILE)
        logger.info(f"Successfully saved cleaned Excel to: {UPDATED_FILE}")
        
    except Exception as e:
        logger.error(f"Error in clean_daily_excel: {str(e)}", exc_info=True)
        raise
