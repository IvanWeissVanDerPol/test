import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

def get_column_index(ws, column_name):
    """Find the column index for a given column name."""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value and str(cell.value).strip() == column_name:
            return col
    return None

def filter_columns(file_path):
    """Create a new file with only the specified columns."""
    # Load the workbook
    wb = load_workbook(filename=file_path)
    ws = wb.active
    
    # List of columns to keep
    columns_to_keep = [
        "columna en informe diario",
        "Hora de Análisis",
        "Saturación (%) (Pureza)",
        "Longitud de onda (nm)",
        "L*",
        "a*",
        "b*",
        "Densidad",
        "% T 550 (2mm)",
        "Semillas L593",
        "Semillas L594",
        "Semillas (0 - 0,5) mm L 593",
        "Semillas (0 - 0,5) mm L 594",
        "Burbujas ( 0,5-1) mm L 593",
        "Burbujas ( 0,5-1) mm L 594",
        "Burbujas ( >1)mm L 593",
        "Burbujas ( >1)mm L 594",
        "Burbujas por Kg - 593",
        "Burbujas por Kg - 594",
        "SiO2",
        "Na2O",
        "CaO",
        "MgO",
        "Al2O3",
        "K2O",
        "SO3",
        "Fe2O3",
        "TiO2",
        "SiO2D (100-S(ox))",
        "Cr2O3",
        "%FeO as Fe2O3",
        "Redox",
        "Viscosidad (°C)",
        "Cooling Time (s)"
    ]
    
    # Create a new workbook for the filtered data
    new_wb = load_workbook(filename=file_path)
    new_ws = new_wb.active
    
    # Get the indices of columns to keep
    columns_indices = []
    for col_name in columns_to_keep:
        col_idx = get_column_index(ws, col_name)
        if col_idx is not None:
            columns_indices.append(col_idx)
        else:
            print(f"Warning: Column '{col_name}' not found in the original file")
    
    # Create a new worksheet for filtered data
    filtered_ws = wb.create_sheet(title="Filtered")
    
    # Copy the specified columns
    for col_idx in columns_indices:
        col_letter = get_column_letter(col_idx)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            filtered_ws.cell(row=row, column=columns_indices.index(col_idx) + 1, value=cell.value)
    
    # Save the filtered file
    original_filename = os.path.basename(file_path)
    filtered_filename = os.path.splitext(original_filename)[0] + '_filtered.xlsx'
    filtered_path = os.path.join(input_dir, filtered_filename)
    
    wb.save(filtered_path)
    print(f"\nFiltered file saved as: {filtered_path}")
    print(f"Total columns kept: {len(columns_indices)}")


