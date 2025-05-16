import os
import logging
from typing import Optional, List
from pathlib import Path
from src.utils.date_utils import format_datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.config.config import COLUMN_CONFIG, FILE_PATHS, DIRECTORIES, VALUES_TO_SET
from src.utils.excel.operations import ExcelOperations
from src.utils.excel_utils import process_dates, save_workbook
from src.exceptions import MissingColumnError

logger = logging.getLogger(__name__)

class ExcelProcessor:
    ESSENTIAL_COLUMNS = ["FECHA", "Hora de Análisis"]
    REQUIRED_COLUMNS = ESSENTIAL_COLUMNS + [
        "Saturación (%) (Pureza)",
        "Longitud de onda (nm)",
        "L*",
        "a*",
        "b*",
        "Densidad",
        "% T 550 (2mm)"
    ]

    def __init__(self, file_path: str):
        self.logger = logger.getChild(self.__class__.__name__)
        self.file_path = file_path
        self.logger.info(f"Initializing with file: {file_path}")
        
        self.wb = load_workbook(filename=file_path, data_only=True)
        self.ws = self.wb.active
        self.logger.debug(f"Workbook loaded with sheets: {self.wb.sheetnames}")

    def __del__(self):
        try:
            self.wb.close()
            self.logger.debug("Workbook closed successfully.")
        except Exception as e:
            self.logger.warning(f"Error closing workbook: {str(e)}")

    def _convert_value(self, value) -> str:
        if value is None:
            return "None"
        if isinstance(value, (int, float)):
            return f"{value} ({type(value).__name__})"
        
        try:
            value_str = str(value)
            if ',' in value_str and '.' not in value_str:
                try:
                    num = float(value_str.replace(',', '.'))
                    return f"{num} (float)"
                except ValueError:
                    pass
            return value_str[:50] + "..." if len(value_str) > 50 else value_str
        except Exception as e:
            self.logger.warning(f"Error converting value to string: {str(e)}")
            return f"ERROR: {str(e)}"

    def _get_column_index(self, column_name: str, ws=None) -> Optional[int]:
        ws = ws or self.ws
        for col in range(1, ws.max_column + 1):
            if str(ws.cell(row=1, column=col).value).strip() == column_name:
                return col
        return None

    def _validate_required_columns(self) -> None:
        self.logger.info("Validating required columns.")
        actual_columns = [cell.value for cell in self.ws[1] if cell.value]

        missing_essential = [col for col in self.ESSENTIAL_COLUMNS if col not in actual_columns]
        if missing_essential:
            raise MissingColumnError(f"Missing essential columns: {', '.join(missing_essential)}")

        missing_optional = [col for col in self.REQUIRED_COLUMNS if col not in actual_columns]
        if missing_optional:
            self.logger.warning(f"Missing optional columns: {missing_optional}")

        self.logger.info("All essential columns are present.")

    def update_dates(self) -> None:
        self.logger.info("Starting date update process.")

        excel_ops = ExcelOperations(self.ws)
        excel_ops.process()
        excel_ops.set_row_values(VALUES_TO_SET, start_row=26)
        excel_ops.move_hora_values()
        process_dates(self.ws)
        excel_ops.delete_rows_and_columns(rows_to_delete=3, cols_to_delete=[6, 5, 4, 3, 1])
        excel_ops.set_column_widths()
        excel_ops.set_row_heights()
        excel_ops.transpose_worksheet(self.wb)

        output_path = os.path.join(DIRECTORIES['output'], f"processed_{os.path.basename(self.file_path)}")
        save_workbook(self.wb, output_path)
        self.file_path = output_path
        self.logger.info(f"Date update completed and saved to: {output_path}")

    def filter_columns(self, sheet_name: str = "Transposed") -> None:
        wb = load_workbook(filename=self.file_path)
        ws = wb[sheet_name]

        columns_to_keep = [
            "Hora de Análisis",
            "Saturación (%) (Pureza)",
            "Longitud  de onda (nm)",
            "L*",
            "a*",
            "b*",
            "Densidad",
            "% T 550 (2mm)",
            "Semillas L593",
            "Semillas L594",
            "Semillas (0 - 0,5) mm L 593",
            "Semillas (0 - 0,5) mm L 594",
            "Burbujas ( 0,5-1) mm L 593",
            "Burbujas ( 0,5-1) mm L 594",
            "Burbujas ( >1)mm L 593",
            "Burbujas ( >1)mm L 594",
            "Burbujas por Kg - 593",
            "Burbujas por Kg - 594",
            "SiO2",
            "Na2O",
            "CaO",
            "MgO",
            "Al2O3",
            "K2O",
            "SO3",
            "Fe2O3",
            "TiO2",
            "SiO2D (100-S(ox))",
            "Cr2O3",
            "%FeO as Fe2O3",
            "Redox",
            "Viscosidad (°C)",
            "Cooling Time (s)"
        ]
        
        filtered_ws = wb.create_sheet(title="Filtered")
        columns_indices = []

        for col_name in columns_to_keep:
            idx = self._get_column_index(col_name, ws)
            if idx:
                columns_indices.append(idx)
            else:
                self.logger.warning(f"Column not found: {col_name}")

        for idx in columns_indices:
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=idx).value
                filtered_ws.cell(row=row, column=columns_indices.index(idx) + 1, value=val)

        
        filtered_path = FILE_PATHS['output']['daily_file_filtered']
        wb.save(filtered_path)

        self.logger.info(f"Filtered file saved as: {filtered_path} with {len(columns_indices)} columns.")

    def transfer_data(self) -> None:
        self.logger.info("Starting data transfer.")
        try:
            input_file = FILE_PATHS['output']['daily_file_filtered']
            if not input_file.exists():
                raise FileNotFoundError(f"Input file not found: {input_file}")

            target_wb = load_workbook(str(input_file))
            target_ws = target_wb.active

            for src_col, tgt_col in COLUMN_CONFIG['column_mapping'].items():
                src_idx = self._get_column_index(src_col)
                tgt_idx = self._get_column_index(tgt_col, target_ws)

                if not src_idx or not tgt_idx:
                    self.logger.warning(f"Column missing in mapping: {src_col} or {tgt_col}")
                    continue

                for row in range(1, self.ws.max_row + 1):
                    val = self.ws.cell(row=row, column=src_idx).value
                    target_ws.cell(row=row, column=tgt_idx).value = val

            output_path = FILE_PATHS['output']['data_ambar_updated']
            target_wb.save(output_path)
            self.logger.info(f"Data transferred successfully to: {output_path}")

        except Exception as e:
            self.logger.error(f"Transfer failed: {str(e)}")
            raise
