from typing import List, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
from .base import ExcelBase
from src.utils.excel_utils import get_column_letter
from src.utils.logging import setup_logging
import pandas as pd

logger = setup_logging()

class ExcelOperations(ExcelBase):
    """Class for general Excel operations."""
    
    def __init__(self, worksheet: Worksheet):
        super().__init__(worksheet)
        self.logger = setup_logging()

    def process(self) -> None:
        """Process the worksheet by applying all operations."""
        self.logger.info("Starting Excel operations processing")
        self.unmerge_cells(column_range=(2, 5))
        self.set_column_widths()
        self.logger.info("Excel operations processing completed")

    def set_row_values(self, values: List[str], start_row: int = 26) -> None:
        """
        Set values in specified rows, handling merged cells.
        
        Args:
            values: List of values to set
            start_row: Row number to start setting values
        """
        logger.info(f"Setting {len(values)} values starting from row {start_row}")
        
        for i, value in enumerate(values, start=start_row):
            self.set_cell_value(i, 2, value)
        
        logger.info("Completed setting row values")

    def move_hora_values(self) -> None:
        """
        Move cells containing "hora" from column F to column B.
        """
        logger.info(f"Starting to move hora values from column F to column B")
        hora_count = 0
        processed_rows = []
        
        # First pass: Identify all hora values and their target cells
        hora_cells = []
        for row in range(1, self.ws.max_row + 1):
            cell_f = self.ws.cell(row=row, column=6)  # Column F
            cell_b = self.ws.cell(row=row, column=2)  # Column B
            
            if cell_f.value and "hora" in str(cell_f.value).lower():
                hora_count += 1
                processed_rows.append(row)
                logger.debug(f"Found hora value in row {row}: {cell_f.value}")
                
                if isinstance(cell_b, MergedCell):
                    logger.debug(f"Target cell in column B is merged, finding top-left cell")
                    top_row, top_col = self.handle_merged_cell(cell_b)
                    target_cell = self.ws.cell(row=top_row, column=top_col)
                    hora_cells.append((cell_f, target_cell))
                else:
                    logger.debug(f"Target cell is direct cell {cell_b.coordinate}")
                    hora_cells.append((cell_f, cell_b))
        
        # Second pass: Move values after identifying all targets
        for source_cell, target_cell in hora_cells:
            logger.debug(f"Moving value from {source_cell.coordinate} to {target_cell.coordinate}")
            target_cell.value = source_cell.value
        
        logger.info(f"Completed moving {hora_count} hora values from column F to B")
        if processed_rows:
            logger.info(f"Processed rows with hora values: {processed_rows}")
        else:
            logger.info("No hora values found in column F")

    def set_column_widths(self) -> None:
        """Set column widths for columns C to the last column."""
        # Start from column C (3) to the last column, adjusting width
        for col in range(1, self.ws.max_column + 1):
            self.ws.column_dimensions[get_column_letter(col)].width = 15

    def set_row_heights(self) -> None:
        """Set row heights for rows 1 to the last row."""
        for row in range(1, self.ws.max_row + 1):
            self.ws.row_dimensions[row].height = 15

    def delete_rows_and_columns(self, rows_to_delete: int = 3, cols_to_delete: List[int] = [6, 5, 4, 3, 1]) -> None:
        """
        Delete specified number of rows and columns from the worksheet.
        
        Args:
            rows_to_delete: Number of rows to delete from the top
            cols_to_delete: List of column indices to delete
        """
        # Delete rows
        for _ in range(rows_to_delete):
            self.ws.delete_rows(1)
        
        # Delete columns in reverse order to avoid index shifting
        for col_idx in sorted(cols_to_delete, reverse=True):
            self.ws.delete_cols(col_idx)

    def transpose_worksheet(self, wb, target_title: str = "Transposed") -> Worksheet:
        """
        Transpose a worksheet and create a new sheet with transposed data.
        
        Args:
            wb: Workbook object
            target_title: Title for the new transposed worksheet
            
        Returns:
            The newly created transposed worksheet
        """
        # Convert to pandas DataFrame and transpose
        df = pd.DataFrame(self.ws.values)
        df_transposed = df.transpose()
        
        # Create a new worksheet and write the transposed data
        ws_new = wb.create_sheet(title=target_title)
        for r_idx, row in enumerate(df_transposed.values, 1):
            for c_idx, value in enumerate(row, 1):
                ws_new.cell(row=r_idx, column=c_idx, value=value)
        
        return ws_new
