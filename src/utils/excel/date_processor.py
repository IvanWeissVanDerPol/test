from typing import List, Optional
from datetime import datetime, time
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
from .base import ExcelBase
from src.exceptions import InvalidTimeFormatError
from src.config.config import COLUMN_CONFIG

class DateProcessor(ExcelBase):
    """Processor for date/time values in Excel worksheets."""
    
    def __init__(self, worksheet: Worksheet):
        super().__init__(worksheet)
        self.date_columns = list(range(7, 200))  # Columns G to HH
        self.default_year = COLUMN_CONFIG.get('default_year', 2025)

    def process(self) -> None:
        """Process date/time values in specified columns."""
        rows_to_update = self._get_rows_to_update()
        self._process_dates(rows_to_update)

    def _get_rows_to_update(self) -> List[int]:
        """Get list of rows to process."""
        return [row for row in range(2, self.ws.max_row + 1) if self._has_time_value(row)]

    def _has_time_value(self, row: int) -> bool:
        """Check if a row has a time value."""
        return any(self.ws.cell(row=row, column=col).value for col in self.date_columns)

    def _process_dates(self, rows_to_update: List[int]) -> None:
        """Process date/time values for each column."""
        for col in self.date_columns:
            self._process_column(col, rows_to_update)

    def _process_column(self, col: int, rows: List[int]) -> None:
        """Process a single column's date/time values."""
        date_cell = self.ws.cell(row=1, column=col)
        date_value = self._parse_date(date_cell.value)
        
        for row in rows:
            time_cell = self.ws.cell(row=row, column=col)
            time_value = self._parse_time(time_cell.value)
            
            if date_value and time_value:
                combined = datetime.combine(date_value, time_value)
                self.set_cell_value(row, col, combined)

    def _parse_date(self, value: any) -> Optional[datetime]:
        """Parse a date value from cell."""
        if isinstance(value, datetime):
            return value.date()
        elif isinstance(value, str):
            try:
                return datetime.strptime(value, '%Y-%m-%d')
            except ValueError:
                return None
        return None

    def _parse_time(self, value: any) -> Optional[time]:
        """Parse a time value from cell."""
        if isinstance(value, time):
            return value
        elif isinstance(value, datetime):
            return value.time()
        elif isinstance(value, str):
            try:
                return datetime.strptime(value, '%H:%M').time()
            except ValueError:
                raise InvalidTimeFormatError(f"Invalid time format: {value}")
        return None
