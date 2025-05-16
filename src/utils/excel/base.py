from abc import ABC, abstractmethod
from typing import Tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
from src.exceptions import ExcelProcessorException

class ExcelBase(ABC):
    """Base class for Excel operations."""
    
    def __init__(self, worksheet: Worksheet):
        self.ws = worksheet
        
    @abstractmethod
    def process(self) -> None:
        """Process the worksheet."""
        pass

    def unmerge_cells(self, column_range: Tuple[int, int]) -> None:
        """Unmerge cells in specified column range."""
        merged_ranges = list(self.ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            if column_range[0] <= merged_range.min_col <= column_range[1]:
                self.ws.unmerge_cells(str(merged_range))

    def handle_merged_cell(self, cell: MergedCell) -> Tuple[int, int]:
        """Get top-left coordinates of a merged cell."""
        for merged_range in self.ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return merged_range.min_row, merged_range.min_col
        raise ExcelProcessorException("Failed to find merged cell range")

    def set_cell_value(self, row: int, col: int, value: any) -> None:
        """Set cell value with merged cell handling."""
        cell = self.ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            top_row, top_col = self.handle_merged_cell(cell)
            target_cell = self.ws.cell(row=top_row, column=top_col)
            target_cell.value = value
        else:
            cell.value = value
