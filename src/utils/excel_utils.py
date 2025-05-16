from typing import List, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from src.exceptions import MissingColumnError, InvalidTimeFormatError, InvalidFilePathError
from src.utils.logging import setup_logging
import pandas as pd
from src.utils.date_utils import format_datetime
from typing import Optional

logger = setup_logging()

def get_column_letter(col_num: int) -> str:
    """Convert a column number to its corresponding Excel column letter."""
    result = []
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result.append(chr(65 + remainder))
    return ''.join(reversed(result))

def save_workbook(wb, output_path: str) -> None:
    """
    Save the workbook with proper error handling and logging.
    
    Args:
        wb: Workbook object to save
        output_path: Path where to save the file
    """
    try:
        wb.save(output_path)
        logger.info(f"File saved successfully to: {output_path}")
    except Exception as e:
        logger.error(f"Error saving file: {str(e)}")
        raise


# Constants for Excel processing
COLUMN_START = 7  # Column G
COLUMN_END = 209  # Column HH
DATE_FORMAT = '%Y-%m-%d'


def process_dates(ws: Worksheet) -> None:
    """
    Process date/time values in specified columns.
    
    Args:
        ws: Excel worksheet
    """
    logger.info("Starting date/time processing...")
    
    # Process columns from G to HH
    for col in range(COLUMN_START, COLUMN_END + 1):  # +1 to include end column
        # Get the date from row 1 of this column
        date_cell = ws.cell(row=1, column=col)
        date_value = date_cell.value
        col_letter = get_column_letter(col)
        
        # Skip if the cell is empty
        if date_value is None:
            continue
            
        try:
            # Process each row that needs updating
            rows_to_update = [4, 23, 34, 55]

            for row in rows_to_update:
                time_cell = ws.cell(row=row, column=col)
                time_value = time_cell.value
                
                if time_value is None:
                    continue
                    
                # Clean the text from the cell before the analysis
                cleaned_time = str(time_value)
                cleaned_time = cleaned_time.replace('hs', '').replace('.', '').replace(' ', '').strip()
                
                # Format the datetime value
                formatted_value = format_datetime(cleaned_time)
                if formatted_value is not None:
                    time_cell.value = formatted_value
                else:
                    logger.warning(f"Row {row}, col {col_letter}: Could not format time: {cleaned_time}")
                    continue
                    
        except Exception as e:
            logger.error(f"Error processing column {col_letter}: {str(e)}")
            continue
    
    logger.info("Date/time processing completed successfully")

def parse_time_string(time_str: str) -> Optional[time]:
    """
    Parse various time formats into a datetime.time object.
    
    Args:
        time_str: Time string to parse
        
    Returns:
        datetime.time object or None if parsing fails
        
    Raises:
        InvalidTimeFormatError: If the time string cannot be parsed
    """
    # Check for empty or NA values
    if not time_str or pd.isna(time_str):
        return None
    
    # Convert to string if it's not already
    time_str = str(time_str).strip()
    
    # Return None for empty strings after stripping
    if not time_str:
        return None
    
    cleaned_time = str(time_str)
    # Clean the time string
    cleaned_time = time_str.lower()  # Convert to lowercase for consistent parsing
    cleaned_time = cleaned_time.replace('hs', '').replace('.', '').replace(' ', '').strip()
    
    # Handle special cases
    if cleaned_time == 'nan':
        return None
    if cleaned_time == '':
        return None
    
    time_formats = [
        '%H:%M',     # 24-hour format
        '%H:%M:%S',  # 24-hour format with seconds
        '%I:%M %p',  # 12-hour format
        '%I:%M%p',   # 12-hour format without space
        '%I.%M %p',  # 12-hour format with period
        '%I.%M%p'    # 12-hour format with period and no space
    ]
    
    for fmt in time_formats:
        try:
            dt = datetime.strptime(cleaned_time, fmt)
            return dt.time()
        except ValueError:
            continue
    raise InvalidTimeFormatError(time_str)

def validate_file_path(path: str) -> str:
    """
    Validate and return absolute path.
    
    Args:
        path: File path to validate
        
    Returns:
        Absolute path to the file
        
    Raises:
        InvalidFilePathError: If the file path is invalid
    """
    try:
        abs_path = Path(path).resolve()
        if not abs_path.exists():
            raise InvalidFilePathError(f"File does not exist: {abs_path}")
        if not abs_path.is_file():
            raise InvalidFilePathError(f"Path is not a file: {abs_path}")
        return str(abs_path)
    except (TypeError, ValueError) as e:
        raise InvalidFilePathError(f"Invalid file path: {str(e)}") from e

def get_column_index(ws, column_name: str) -> Optional[int]:
    """
    Find the column index for a given column name.
    
    Args:
        ws: Excel worksheet
        column_name: Name of the column to find
        
    Returns:
        Column index (1-based) or None if not found
        
    Raises:
        MissingColumnError: If the column is required but not found
    """
    # Try different variations of the column name
    variations = [
        column_name,
        column_name.lower(),
        column_name.upper(),
        column_name.strip(),
        column_name.replace(' ', '_'),
        column_name.replace('_', ' '),
        column_name.replace(' ', ''),  # Remove all spaces
        column_name.replace('-', ''),  # Remove hyphens
        column_name.replace('(', ''),  # Remove parentheses
        column_name.replace(')', '')   # Remove parentheses
    ]
    
    # Also try variations with common typos
    typo_variations = [
        column_name.replace('ó', 'o'),  # Replace accented o
        column_name.replace('í', 'i'),  # Replace accented i
        column_name.replace('á', 'a'),  # Replace accented a
        column_name.replace('ú', 'u'),  # Replace accented u
        column_name.replace('é', 'e'),  # Replace accented e
        column_name.replace('ñ', 'n')   # Replace ñ
    ]
    
    variations.extend(typo_variations)
    
    actual_columns = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            cell_value = str(cell.value).strip()
            actual_columns.append(cell_value)
            # Try all variations
            for variation in variations:
                if variation.lower() == cell_value.lower():
                    return col
    
    # Log the actual column headers we found
    logger.debug(f"Actual column headers found: {', '.join(actual_columns)}")
    
    # Check if this is a required column
    required_columns = COLUMN_CONFIG['columns_to_keep']
    if column_name in required_columns:
        # Log the actual column headers we found
        actual_columns = [str(ws.cell(row=1, column=c).value).strip() 
                         for c in range(1, ws.max_column + 1) 
                         if ws.cell(row=1, column=c).value]
        raise MissingColumnError(f"Required column not found: {column_name}. Found columns: {', '.join(actual_columns)}")
    
    return None
