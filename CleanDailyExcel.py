import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from datetime import datetime, time
import os
from typing import List, Optional

def get_column_letter(col_num: int) -> str:
    """Convert a column number to its corresponding Excel column letter."""
    result = []
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result.append(chr(65 + remainder))
    return ''.join(reversed(result))

def unmerge_columns(ws) -> None:
    """Unmerge cells in the specific range B26 to E31."""
    print("\nUnmerging cells in range B26 to E31...")
    # Define the range to unmerge
    start_row = 26
    end_row = 31
    start_col = 2  # Column B
    end_col = 5    # Column E
    
    # Unmerge any merged cells in the specified range
    for merged_range in list(ws.merged_cells.ranges):
        # Check if the merged range overlaps with our target range
        if (start_row <= merged_range.min_row <= end_row and
            start_row <= merged_range.max_row <= end_row and
            start_col <= merged_range.min_col <= end_col and
            start_col <= merged_range.max_col <= end_col):
            ws.unmerge_cells(str(merged_range))
            print(f"Unmerged cells in range: {merged_range}")
    
    # Also unmerge any cells in columns A-G that might be merged
    print("\nUnmerging any remaining merged cells in columns A-G...")
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_col <= 7:  # Column G is 7
            ws.unmerge_cells(str(merged_range))
            print(f"Unmerged cells in range: {merged_range}")

def set_row_values(ws, values: List[str], start_row: int = 26) -> None:
    """Set values in specified rows, handling merged cells."""
    print("\nSetting values in rows 26-33...")
    for i, value in enumerate(values, start=start_row):
        cell = ws.cell(row=i, column=2)
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    top_left.value = value
                    print(f"Setting merged cell value in row {i}: {value}")
                    break
        else:
            cell.value = value
            print(f"Setting value in row {i}: {value}")

def move_hora_values(ws) -> None:
    """Move cells containing "hora" from column F to column B."""
    print("\nMoving hora values from column F to column B...")
    for row in range(1, ws.max_row + 1):
        cell_f = ws.cell(row=row, column=6)  # Column F
        cell_b = ws.cell(row=row, column=2)  # Column B
        
        if cell_f.value and "hora" in str(cell_f.value).lower():
            if isinstance(cell_b, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell_b.coordinate in merged_range:
                        top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                        top_left.value = cell_f.value
                        print(f"Moving hora value to merged cell in row {row}")
                        break
            else:
                cell_b.value = cell_f.value
                print(f"Moving hora value to row {row}")
            cell_f.value = None

def parse_time_string(time_str: str) -> Optional[time]:
    """Parse various time formats into a datetime.time object."""
    # Clean up the time string
    cleaned_time = str(time_str)
    cleaned_time = cleaned_time.replace('hs', '').replace('.', '').strip()
    cleaned_time = cleaned_time.replace(' ', '')  # Remove spaces
    
    # Try multiple time formats
    time_formats = [
        '%H:%M',     # 24-hour format
        '%H:%M:%S',  # 24-hour format with seconds
        '%I:%M %p',  # 12-hour format
        '%I:%M%p',   # 12-hour format without space
        '%I.%M %p',  # 12-hour format with period
        '%I.%M%p'    # 12-hour format with period and no space
    ]
    
    for fmt in time_formats:
        try:
            return datetime.strptime(cleaned_time, fmt).time()
        except ValueError:
            continue
    return None

def process_dates(ws, rows_to_update: List[int]) -> None:
    """Process date/time values in columns G to HH."""
    print("\nStarting date/time processing...")
    
    # Process each column from 7 to 209 (columns 7 to 209)
    for col in range(7, 209):
        # Get the date from row 1 of this column
        date_cell = ws.cell(row=1, column=col)
        date_value = date_cell.value
        
        # Skip if the cell is empty
        if date_value is None:
            col_letter = get_column_letter(col)
            print(f"Column {col_letter}: Empty date cell, skipping")
            continue
            
        try:
            # Convert date string to datetime object
            if isinstance(date_value, str):
                day, month, year = date_value.split('-')
                year = f"20{year}"
                date_obj = datetime.strptime(f"{day}-{month}-{year}", '%d-%b-%Y')
            else:
                date_obj = date_value
            
            col_letter = get_column_letter(col)
            print(f"\nColumn {col_letter}: Base date: {date_obj}")
            
            # Process each row that needs updating
            for row in rows_to_update:
                time_cell = ws.cell(row=row, column=col)
                time_value = time_cell.value
                
                if time_value is None:
                    print(f"Row {row}, col {col_letter}: Empty time cell, skipping")
                    continue
                    
                if isinstance(time_value, time):
                    time_value = time_value.strftime('%H:%M:%S')
                elif not isinstance(time_value, str):
                    print(f"Row {row}, col {col_letter}: Invalid time format: {type(time_value)}")
                    continue
                    
                try:
                    print(f"\nProcessing row {row}, col {col_letter}")
                    print(f"Original time: {time_value}")
                    
                    time_obj = parse_time_string(time_value)
                    if time_obj is None:
                        print(f"Could not parse time: {time_value}")
                        continue
                    
                    print(f"Parsed time: {time_obj}")
                    combined_datetime = datetime.combine(date_obj.date(), time_obj)
                    print(f"Combined datetime: {combined_datetime}")
                    
                    time_cell.value = combined_datetime
                    print(f"Updated row {row}, col {col} with: {combined_datetime}")
                except Exception as e:
                    print(f"Error processing row {row}, col {col}: {str(e)}")
                    continue
        except Exception as e:
            col_letter = get_column_letter(col)
            print(f"Error converting base date in column {col_letter}: {str(e)}")
            continue

def set_column_widths(ws) -> None:
    """Set column widths for columns C to the last column."""
    # Start from column C (3) to the last column
    for col in range(3, 209):
        ws.column_dimensions[get_column_letter(col)].width = 15

def update_dates_in_excel(file_path: str) -> None:
    """Main function to process the Excel file."""
    print(f"\nProcessing Excel file: {file_path}")
    
    # Load the workbook
    wb = load_workbook(filename=file_path)
    ws = wb.active
    
    # Set values in rows 26-33
    values_to_set = [
        "Semillas L593",
        "Semillas L594",
        "Semillas (0 - 0,5) mm L 593",
        "Semillas (0 - 0,5) mm L 594",
        "Burbujas ( 0,5-1) mm L 593",
        "Burbujas ( 0,5-1) mm L 594",
        "Burbujas ( >1)mm L 593",
        "Burbujas ( >1)mm L 594"
    ]
    
    # Unmerge cells in columns A-G
    unmerge_columns(ws)
    
    # Set values and move hora values
    set_row_values(ws, values_to_set)
    move_hora_values(ws)
    
    # Process date/time values
    rows_to_update = [4, 23, 34, 55]
    process_dates(ws, rows_to_update)
    
    # Set column widths for columns C to the last column
    set_column_widths(ws)
    
    # Save the updated file
    original_filename = os.path.basename(file_path)
    updated_filename = os.path.splitext(original_filename)[0] + '_updated.xlsx'
    input_dir = os.path.dirname(file_path)
    data_dir = os.path.join(input_dir, 'data')
    os.makedirs(data_dir, exist_ok=True)
    updated_path = os.path.join(data_dir, updated_filename)
    
    # Remove the first row 3 times
    for row in range(1, 4):
        ws.delete_rows(1)
    ws.delete_cols(6)
    ws.delete_cols(5)
    ws.delete_cols(4)
    ws.delete_cols(3)
    ws.delete_cols(1)
    
    # Convert to pandas DataFrame and transpose
    df = pd.DataFrame(ws.values)
    df_transposed = df.transpose()
    
    # Create a new worksheet and write the transposed data
    ws_new = wb.create_sheet(title="Transposed")
    for r_idx, row in enumerate(df_transposed.values, 1):
        for c_idx, value in enumerate(row, 1):
            ws_new.cell(row=r_idx, column=c_idx, value=value)
    
    # Remove the original sheet
    del wb[ws.title]
    
    # Rename the new sheet to the original name
    ws_new.title = ws.title
    
    wb.save(updated_path)
    print("\nProcessing complete!")
    print(f"Updated file saved as: {updated_path}")
    print("All formatting (colors, column widths, row heights) has been preserved.")
    print(f"Total columns processed: {ws.max_column}")

if __name__ == "__main__":
    file_path = r"c:\Users\weiss\Desktop\test\data\A-INFORME QUÍMICO DIARIO 2025 Macro prueba.xlsx"
    update_dates_in_excel(file_path)
